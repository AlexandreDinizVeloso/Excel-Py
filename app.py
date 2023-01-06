#Importação
from PySimpleGUI import PySimpleGUI as sg
import openpyxl
import os.path
from os import path

#Definição do menu
def Menu(nome_tela, layout):
    janelaM = sg.Window(nome_tela, layout = layout, finalize = True)
    while True:
        evento, valores = janelaM.read()
        if evento == sg.WIN_CLOSED:
            break

layoutM = [
    [sg.Text("O que deseja fazer?")],
    [sg.Button("Planilha não existente")],
    [sg.Button("Adicionar medicamento"), sg.Button("Ver medicamentos")]
]

#Interface
janela = sg.Window("Menu", layoutM)

#Eventos
while True:
    evento, valores = janela.read()
    if evento == sg.WIN_CLOSED:
        break
    elif evento == "Planilha não existente":
        #Verificação da existência da planilha
        if path.exists("Planilha Remédios.xlsx") == False:
            planilha = openpyxl.Workbook()

            #Criação da página da planilha
            rem_plan = planilha["Sheet"]
            rem_plan.title = "Remédios"

            #Seleção de página
            rem_plan = planilha["Remédios"]
            rem_plan.append(["Medicamento", "Data de Entrada", "Gramatura", "Quantidade", "Validade"])

            #Salvar planilha
            planilha.save("Planilha Remédios.xlsx")

            #Execução gráfica
            janela.hide()
            layoutC = [
                [sg.Text("Planilha criada.")]
            ]
            Menu("Criação", layoutC)
            janela.un_hide()
        else:
            janela.hide()
            layoutC = [
                [sg.Text("Planilha já existente.")],
            ]
            Menu("Criação", layoutC)
            janela.un_hide()

    elif evento == "Adicionar medicamento":
        #Carregar planilha
        if path.exists("Planilha Remédios.xlsx") == False:
            janela.hide()
            layoutA = [
                [sg.Text("Planilha não encontrada.")]
            ]
            Menu("Adição", layoutA)
            janela.un_hide()
        else:
            planilha = openpyxl.load_workbook("Planilha Remédios.xlsx")

            #Seleção de página
            rem_plan = planilha["Remédios"]

            #Adição de uma linha na planilha
            janela.hide()
            layoutA = [
                [sg.Text("Adição de medicação")],
                [sg.Text("Medicação"), sg.Input(key = "medicação")],
                [sg.Text("Data de Entrada"), sg.Input(key = 'entrada')],
                [sg.Text("Gramatura"), sg.Input(key = "gramatura")],
                [sg.Text("Quantidade"), sg.Input(key = "quantidade")],
                [sg.Text("Validade"), sg.Input(key = "validade")],
                [sg.Button("Salvar")]
            ]
            janelaA = sg.Window("Menu", layoutA)
            evento, valores = janelaA.read()
            while True:
                if evento == sg.WIN_CLOSED:
                    break
                if evento == "Salvar":
                    rem_plan.append([valores["medicação"], valores["entrada"], valores["gramatura"], valores["quantidade"], valores["validade"]])
                    #Salvar documento
                    planilha.save("Planilha Remédios.xlsx")
                    janelaA.close()
                    layoutA = [
                        [sg.Text("Medicação adicionada.")]
                    ]
                    Menu("Adição", layoutA)
                    break
            janela.un_hide()
    elif evento == "Ver medicamentos":
        #Carregar planilha
        if path.exists("Planilha Remédios.xlsx") == False:
            janela.hide()
            layoutA = [
                [sg.Text("Planilha não encontrada.")]
            ]
            Menu("Adição", layoutA)
            janela.un_hide()
        else:
            planilha = openpyxl.load_workbook("Planilha Remédios.xlsx")

            #Seleção de página
            rem_plan = planilha["Remédios"]

            #Visualização de linhas da planilha
            lista = []
            for rows in rem_plan.iter_rows():
                lista.append([rows[0].value, rows[1].value, rows[2].value, rows[3].value, rows[4].value])
            print(lista)
            s = '\n'.join([str(i) for i in lista])
            sg.PopupScrolled(*lista, title="Lista de Medicamentos")